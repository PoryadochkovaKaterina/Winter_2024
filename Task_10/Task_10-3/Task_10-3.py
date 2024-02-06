import openpyxl
import statistics
from openpyxl import Workbook
wb = Workbook()
wb.save('test1.xlsx')

wb = openpyxl.load_workbook('test1.xlsx')   # Открытие книги
# print(wb.sheetnames)   # Список листов
wb.create_sheet('Stat')
wb.save('test1.xlsx')   # Сохранение файла
# print(wb.sheetnames)

ws = wb.active
ws['A1'].value = 'Котов'
ws['A2'].value = 'Федоров'
ws['A3'].value = 'Сидоров'
ws['A4'].value = 'Иванов'
ws['A5'].value = 'Петров'
ws['B1'].value = '500'
ws['B2'].value = '150'
ws['B3'].value = '200'
ws['B4'].value = '400'
ws['B5'].value = '350'
#print(ws.max_row, ws.max_column)
wb.save('test1.xlsx')

lst = []
res = []
for i in range(1, ws.max_row +1):   # Печать колонки номер 2
    lst.append(ws.cell(row=i, column=2).value)

res.append(max(lst))
res.append(min(lst))
res.append(sum(list(map(int, lst))))
res.append(sum(list(map(int, lst)))/len(lst))
res.append(statistics.median(lst))
print(lst, res)

wb.active = wb['Stat']
ws_S = wb.active
ws_S['A1'].value = 'Максимальное значение'
ws_S['A2'].value = 'Минимальное значение'
ws_S['A3'].value = 'Сумма'
ws_S['A4'].value = 'Среднее арифметическое'
ws_S['A5'].value = 'Медиана'
wb.save('test1.xlsx')

for i in range(1, ws.max_row +1):
    ws_S.cell(row=i, column=2).value = res[i-1]
wb.save('test1.xlsx')

for i in range(1, ws.max_row +1):
    for j in range(1, ws.max_column +1):
        print(ws_S.cell(row=i, column=j).value, end='\t')
    print()





