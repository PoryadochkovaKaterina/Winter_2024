import openpyxl
from openpyxl import Workbook
wb = Workbook()
wb.save('test1.xlsx')

wb = openpyxl.load_workbook('test1.xlsx')   # Открытие книги
# print(wb.sheetnames)   # Список листов
wb.create_sheet('New')
wb.save('test1.xlsx')   # Сохранение файла
# print(wb.sheetnames)

ws = wb.active
ws['A1'].value = 'Котов'
ws['A2'].value = 'Федоров'
ws['A3'].value = 'Сидоров'
ws['A4'].value = 'Иванов'
ws['A5'].value = 'Петров'
ws['B1'].value = '500'
ws['B2'].value = '150'
ws['B3'].value = '200'
ws['B4'].value = '400'
ws['B5'].value = '350'
#print(ws.max_row, ws.max_column)
wb.save('test1.xlsx')
lst = []
for s in range(1, ws.max_row+1):
    lst.append(ws.cell(row=s, column=2).value)
    su = sum(list(map(int, lst)))

li = sorted(ws.values, key=lambda col: [col[1], col], reverse=True)

wb.active = wb['New']
ws_S = wb.active
for row in li:
    ws_S.append(row)
dl1 = ws.max_row + 1
ws_S.cell(row=dl1, column=1).value = 'Итого'
ws_S.cell(row=dl1, column=2).value = su

for i in range(1, ws_S.max_row + 1):
    for j in range(1, ws_S.max_column + 1):
        print(ws_S.cell(row=i, column=j).value, end='\t')
    print()

wb.save('test1.xlsx')