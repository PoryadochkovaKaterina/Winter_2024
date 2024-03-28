# dct = dict.fromkeys([0,-1,1,0,-1,0,-1,1], 0)
# print(len(dct))
# print(sorted(dct, key=lambda x: -abs(x)))

# tcid = {1: {1: {1: {1: 1111}}}}
# print(len(tcid))
# print(tcid[1])
# print(tcid[1][1][1][1])

# Задание 9-1
# gens = {'G': 'C', 'C': 'G', 'T': 'A', 'A': 'T'}
# s = input()
# res = ''
# for i in s:
#     # res += gens[i]
#     res += gens.get(i, '*')   # В случае, если указана неверная буква, которой нет в словаре
# print(res)

# Задание 9-2
# vo = 'ауоыэияюёе'
# lst = ['кино', 'питон', "дорога", "поросенок", "итог", "титан", "погост", "идея"]
# sample = 'питон'
#
# def sim(st):
#     res = []
#     for i in range(len(st)):
#         if st[i] in vo:
#             res.append(i)
#     return res
# # print(sim('поросенок'))
# sample_sim = sim(sample)
# for i in lst:
#     if sim(i) == sample_sim:
#         print(i)

# Задание 9-3
# s = input().lower()
# dct = {}
# for i in s:
#     dct[i] = dct.get(i, 0) + 1    # Замена большого количества if
# lst = sorted(dct, key=lambda x: (-dct[x], x))
# for i in lst[:10]:
#     print(f'{i} - {dct[i]}')

# f = open('File/text.txt', encoding='utf-8')
# s = f.read()   # Прочитал
# print(s)    # Напечатал
# f.close()   # Закрыл

# 'r' - чтение файла
# 'w' - запись в файле
# 'x' - эклюзивное создание
# 'a' - добавление информации в конец файла
# '+' - открытие для чтения и записи

# f = open('File/text.txt', 'r', encoding='utf-8')   # read стоит по умолчанию, но лучше указывать, что для чтения
# s = f.read()   # Прочитал
# print(s)    # Напечатал
# f.close()   # Закрыл

# fp = open(file, mode, encoding)

# f = open('File/text.txt', 'r')
# print(f.read(10))   # Close Your (10 первых символов из файла. Пробел и перевод как один символ
# print(f.read(4))    # Doo (следуюеие 4 символа из файла. Пробел и перевод как один символ
# print(f.read())
# f.close()

# f = open('File/text.txt', 'r')
# s = f.readlines()   # Каждая строка файла с элементами. Удобнее
# for i in range(len(s)):
#     print(i, len(s[i]), s[i].strip())
# f.close()

# Печать с помощью функции и номером строк
# f = open('File/text.txt', 'r')
# # print(f.readline().strip())   # Чтение только первой строки
# # print(f.readline().strip())
# for i in range(10):
#     print(i + 1, f.readline().strip())
# f.close()

# Итерирование файла
# f = open('File/text.txt')
# for line in f:
#     print(line.strip())
# for l, k in enumerate(f):
#     print(l, k.strip())
# f.close()

# Запись в файл
# f = open('File/text.txt', 'w', encoding='utf-8')
# f.write('Чтобы что-то да было\n')
# f.write('А может и не было')
# f.close()

# f = open('File/text.txt', 'r', encoding='utf-8')
# for i in f:
#     print(i.strip())
# f.close()

# f = open('File/text.txt', 'a', encoding='utf-8')
# f.write('Что-то хотелось еще\n')
# f.write('Да не будет')
# f.close()
#
# f = open('File/text.txt', 'r', encoding='utf-8')
# for i in f:
#     print(i.strip())
# f.close()

# f = open('File/text.txt', 'w', encoding='utf-8')
# line = ['Чтобы что-то было\n',
#         'А может и не было\n',
#         'Зачем-то, но должно быть\n',]
# f.writelines(line)
# f.close()
#
# f = open('File/text.txt', 'r', encoding='utf-8')
# for i in f:
#     print(i.strip())
# f.close()

# 1. Открыть файл
# 2. Считать информацию в конструкцию (словарь, список, кортеж и т.д.)
# 3. Обработать информацию, напечатать результат, ввести результат в этот или другой файл
# 4. Закрыть файл

# f_in = open('File/text.txt', 'r', encoding='utf-8')
# s = f_in.readlines()
# f_in.close()
#
# f_out = open('File/text1.txt', 'w', encoding='utf-8')
# for i in s:
#     for j in i:
#         if j.isdigit():
#             f_out.writelines(i)
#             break
# f_out.close()
#
# f = open('File/text1.txt', 'r', encoding='utf-8')
# for i in f:
#     print(i)
# f.close()

# f_in = open('File/text.txt', 'r', encoding='utf-8')
# s = f_in.read()
# f_in.close()
#
# f_out = open('File/text1.txt', 'w', encoding='utf-8')
# for i in s[::10]:   # Печать каждого второго знака из одного файла
#     f_out.write(i)
# f_out.close()
#
# f = open('File/text1.txt', 'r', encoding='utf-8')
# for i in f:
#     print(i)
# f.close()

# f = open('File/text.txt', 'w', encoding='utf-8')
# print(*object, sep='', end='\n', file=f)


# f_in = open('File/text.txt', 'r', encoding='utf-8')
# s = f_in.readlines()
# f_in.close()
#
# f_out = open('File/text1.txt', 'w', encoding='utf-8')
# for i in s:   # Печать каждого второго знака из одного файла
#     print(i, file = f_out, end = '')
#     print(i.strip())
# f_out.close()
#
# f = open('File/text1.txt', 'r', encoding='utf-8')
# for i in f:
#     print(i)
# f.close()

# Запись с возможность чтения
# 'w+'
# 'r+'
# При w+ полностью удаляется вся информация

# f_in = open('File/text.txt', 'r', encoding='utf-8')
# s = f_in.readlines()
# f_in.close()
#
# f_out = open('File/text1.txt', 'at', encoding='utf-8')
# for i in s:   # Печать каждого второго знака из одного файла
#     print(i, file = f_out, end = '')
#     print(i.strip())
# f_out.close()
#
# f = open('File/text1.txt', 'r', encoding='utf-8')
# for i in f:
#     print(i)
# f.close()

# Если close не писать, то можно использовать with

# with open('File/text.txt', 'r', encoding='utf-8') as f_in:
#     s = f_in.readlines()
#     print(s)

# Задание
# s = ['это первая строка\n',
#      'это вторая строка\n',
#      'А это третья строка\n']
# with open('File/text.txt', 'w') as f:
#     f.writelines(s)
#
# with open('File/text.txt') as f2:
#     s = f2.readlines()
#     print(s)
#     with open('File/text1.txt', 'w') as f3:
#         for i in s:
#             w = ' '.join(sorted(i.split()))
#             print(w, file=f3)
#
# with open('File/text1.txt') as f4:
#     z = f4.readlines()
#     print(z)


# EXCEL
# Функции openpyxl
# - Книга
# - Страница
# - Ячейка

# Создание нового файла
# import openpyxl
# from openpyxl import Workbook
# wb = Workbook()
# wb.save('File/test.xlsx')

# import openpyxl
# wb = openpyxl.load_workbook('File/test.xlsx')   # Открытие книги
# # print(wb.sheetnames)   # Список листов
#
# ws = wb.active   # Какой активный лист
# # print(ws.title)   # Каков его title
#
# # wb.create_sheet('New')   # Создание нового листа
# # wb.create_sheet('New1')
# # ws_n = wb['New1']
# # wb.remove(ws_n)
# # wb.save('File/test.xlsx')   # Сохранение файла
# print(wb.sheetnames)   # Просмотр списков листов


# Запись данных в файл
# import openpyxl
# wb = openpyxl.load_workbook('File/test.xlsx')
# ws = wb.active
# ws['A1'].value = 213
# ws['B2'].value = 'ЯчейкаB2'
# print(ws['A1'].value)
# print(ws['B2'].value)
# wb.save('File/test.xlsx')   # Сохранение файла


# Максимальная сторока и колонка
import openpyxl
wb = openpyxl.load_workbook('../File/test.xlsx')
ws = wb.active
ws['A1'].value = 213
ws['B2'].value = 'ЯчейкаB2'
# print(ws['A1'].value)
# print(ws['B2'].value)
# print(ws.max_row, ws.max_column)

# for i in range(ws.max_row):
#     for j in range(ws.max_row):
#         ws.cell(row=i + 1, column=j + 1).value = i * 10 + j
#     print()
#
# for i in range(ws.max_row):
#     for j in range(ws.max_row):
#         print(i+1, j+1, ws.cell(row=i + 1, column=j + 1).value)
#     print()

# for i in range(1, ws.max_row):   # Печать колонки номер 2
#     print(i, ws.cell(row=i, column=2).value)
# wb.save('File/test.xlsx')   # Сохранение файла