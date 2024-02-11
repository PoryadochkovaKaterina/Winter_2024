import openpyxl
import statistics
from openpyxl import Workbook
wb = Workbook()
wb.save('text1.xlsx')
wb = openpyxl.load_workbook('text1.xlsx')
ws = wb.active
wb.save('text1.xlsx')

with open('text1.txt', encoding='utf-8') as ku:
    s = ku.readlines()
    lst = []
    for i in s:
        lst.append(i.strip().split(','))
    # print(lst)

wb.active = wb['Sheet']
ws = wb.active
lst_sor = sorted(lst, key=lambda x: (x[3], x[1], x[2]))
# print(lst_sor)
su = 0
for row, col in enumerate(lst_sor, 1):
    for k in range(5):
        ws.cell(row, k + 1).value = col[k]
    su += int(col[4])
dl1 = ws.max_row + 1
ws.cell(dl1, 4).value = 'Итого: '
ws.cell(dl1, 5).value = su
wb.save('text1.xlsx')





