import openpyxl
import statistics
from openpyxl import Workbook
wb = Workbook()
wb.save('text.xlsx')
wb = openpyxl.load_workbook('text.xlsx')
ws = wb.active
wb.save('text.xlsx')

with open('text.txt', encoding='utf-8') as ku:
    s = ku.readlines()
    lst = []
    for i in range(9):
        for item in s:
            lst.append(item.strip().split(', ')[i])
        lst_r = tuple(lst)
    size = 5
    lst_re = [lst_r[i:i + size] for i in range(0, len(lst_r), size)]
    print(lst_re)

wb.active = wb['Sheet']
ws = wb.active
lst_sor = sorted(lst_re, key=lambda x: (x[3], x[1], x[2]))
print(lst_sor)
su = 0
for row, col in enumerate(lst_sor, 1):
    for k in range(5):
        ws.cell(row, k + 1).value = col[k]
    su += int(col[4])
dl1 = ws.max_row + 1
ws.cell(dl1, 4).value = 'Итого: '
ws.cell(dl1, 5).value = su
wb.save('text.xlsx')





